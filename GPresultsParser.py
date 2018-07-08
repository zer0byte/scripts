#!/usr/bin/python
# GPresults Parser v2.0
# By Zer0byte

# modules in standard library
import bs4
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Fill


#gpresult = sys.argv[1]
#url = gpresult
#page = open(url)
try:
    page = open(str(sys.argv[1]))
except Exception, e:
        print " "
        print " Useage gpresult.py <Group policy HTML file>"
        print " "
        sys.exit()

soup = bs4.BeautifulSoup(page, 'html.parser')
FileName = soup.find("td", text="Computer name").find_next_sibling("td").text
a, b = FileName.split("\\")
def StartUp():
    os.system("clear")
    print "-" * 80
    print "Group Polciy Results Parser by: Bilal Bokhari"
    print " "
    print " "
    print "Parsing group policy results for computer name: " + b
    print "-" * 80

StartUp()
wb = Workbook()
# grab the active worksheet
ws = wb.active
# 1st Heading
ws.font = Font(bold=True)
ws['A1'] = "#"
ws['A1'].font = Font(bold=True)
# 2nd Heading
ws['B1'] = "Finding"
ws['B1'].font = Font(bold=True)
# 3rd Heading
ws['C1'] = b
ws['C1'].font = Font(bold=True)

configs = {
    1:'Minimum password length',
    2:'Enforce password history',
    3:"Network access: Do not allow anonymous enumeration of SAM accounts and shares",
    4:"Act as part of the operating system",
    5:"Add workstations to domain",
    6:"Allow log on through Terminal Services",
    7:"Do not allow password to be saved",
    8:"Microsoft network server: Digitally sign communications (always)",
    9:"Microsoft network server: Digitally sign communications (if client agrees)",
    10:"Domain controller: LDAP server signing requirements",
    11:"Interactive logon: Require Domain Controller authentication to unlock workstation",
    12:"Interactive logon: Do not display last user name",
    13:"Interactive logon: Message title for users attempting to log on",
    14:"Network security: LAN Manager authentication level",
    15:"Network access: Shares that can be accessed anonymously",
    16:"User Account Control: Switch to the secure desktop when prompting for elevation",
    17:"Maximum application log size",
    18:"Maximum security log size",
    19:"Maximum system log size"
}

for i, v in configs.items():
    ws['A'+str(i+1)] = i
    ws['B'+str(i+1)] = str(v)
    try:
        configLookUp = soup.find("td", text=v).find_next_sibling("td").text
        print "["+str(i) + "] " + v + " = " + str(configLookUp)
        ws['C' + str(i + 1)] = configLookUp
    except Exception, e:
        configLookUp = "Not configured / Key not found"
        print "[" + str(i) + "] " + v + " = " + str(configLookUp)
        ws['C' + str(i + 1)] = configLookUp

wb.save(str(b) + ".xlsx")

print " "
print "[+] Saving parsed results in %s.xlsx" % b
print " "
print "[+] Parsing complete!"
print " "

sys.exit()
